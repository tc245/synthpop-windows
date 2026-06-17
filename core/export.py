"""PDF and Excel export for the synthetic data quality report."""
import base64
import io


# ── PDF ───────────────────────────────────────────────────────────────────────

def _render_pdf_html(report: dict) -> str:
    """Generate Qt-safe HTML for QPrinter / QTextDocument.

    Rules enforced here:
    - No <thead>/<tbody>/<tfoot>  — Qt ignores unknown tags AND their children
    - No CSS class selectors      — QTextDocument support is unreliable
    - No border-collapse CSS      — use cellspacing/cellpadding attributes instead
    - No nested tables            — outer layout is headings, not a wrapper table
    - No <code> tags              — plain text only
    """
    _TH = "padding:4px 8px;background:#f0f0f0;font-weight:bold;border:1px solid #ccc;"
    _TD = "padding:4px 8px;border:1px solid #ccc;"
    _H2 = "font-size:15px;color:#333;border-bottom:2px solid #9063CD;padding-bottom:3px;margin-top:14px;"
    _H3 = "font-size:13px;color:#555;margin-top:8px;margin-bottom:2px;"
    _BODY = "font-family:Arial,sans-serif;font-size:13px;margin:8px;"

    def _v(val):
        return "—" if val is None else str(val)

    def _pass(passed):
        if passed is None:
            return f"<td style='{_TD}text-align:center;'>—</td>"
        if passed:
            return f"<td style='{_TD}text-align:center;color:green;font-weight:bold;'>✓</td>"
        return f"<td style='{_TD}text-align:center;color:red;font-weight:bold;'>✗</td>"

    p = [f"<html><body style='{_BODY}'>"]
    p.append(
        f"<p style='color:#555;'>Original: <b>{report['n_orig']:,}</b> rows"
        f"&nbsp;·&nbsp;Synthetic: <b>{report['n_synth']:,}</b> rows</p>"
    )

    # ── Numeric summary ───────────────────────────────────────────────────────
    ns = report.get("numeric_summary")
    if ns and ns["rows"]:
        p.append(f"<h2 style='{_H2}'>Numeric Variable Summary</h2>")
        p.append(f"<h3 style='{_H3}'>Summary statistics</h3>")
        p.append("<table cellspacing='0' cellpadding='0' width='100%'>")
        # Header — no colspan; write var name once then empty header for synth col
        p.append(f"<tr><th style='{_TH}'>Statistic</th>")
        for col in ns["num_vars"]:
            p.append(f"<th style='{_TH}'>{col} (Real)</th>")
            p.append(f"<th style='{_TH}'>{col} (Synth)</th>")
        p.append("</tr>")
        for row in ns["rows"]:
            p.append(f"<tr><td style='{_TD}'>{row['stat']}</td>")
            for col in ns["num_vars"]:
                p.append(f"<td style='{_TD}text-align:right;'>{_v(row.get(col+'__real'))}</td>")
                p.append(f"<td style='{_TD}text-align:right;'>{_v(row.get(col+'__synth'))}</td>")
            p.append("</tr>")
        p.append("</table>")

    if report.get("ks_tests"):
        p.append(f"<h3 style='{_H3}'>Kolmogorov–Smirnov tests</h3>")
        p.append("<table cellspacing='0' cellpadding='0'>")
        p.append(
            f"<tr><th style='{_TH}'>Variable</th>"
            f"<th style='{_TH}'>KS statistic</th>"
            f"<th style='{_TH}'>p-value</th>"
            f"<th style='{_TH}'>Pass</th></tr>"
        )
        for row in report["ks_tests"]:
            p.append(
                f"<tr><td style='{_TD}'>{row['variable']}</td>"
                f"<td style='{_TD}text-align:right;'>{_v(row['statistic'])}</td>"
                f"<td style='{_TD}text-align:right;'>{_v(row['pvalue'])}</td>"
                f"{_pass(row['pass'])}</tr>"
            )
        p.append("</table>")

    # ── Categorical summary ───────────────────────────────────────────────────
    if report.get("chi2_tests") or report.get("cat_freq"):
        p.append(f"<h2 style='{_H2}'>Categorical Variable Summary</h2>")

    if report.get("chi2_tests"):
        p.append(f"<h3 style='{_H3}'>Chi-squared tests</h3>")
        p.append("<table cellspacing='0' cellpadding='0'>")
        p.append(
            f"<tr><th style='{_TH}'>Variable</th>"
            f"<th style='{_TH}'>χ² statistic</th>"
            f"<th style='{_TH}'>p-value</th>"
            f"<th style='{_TH}'>Pass</th></tr>"
        )
        for row in report["chi2_tests"]:
            p.append(
                f"<tr><td style='{_TD}'>{row['variable']}</td>"
                f"<td style='{_TD}text-align:right;'>{_v(row['statistic'])}</td>"
                f"<td style='{_TD}text-align:right;'>{_v(row['pvalue'])}</td>"
                f"{_pass(row['pass'])}</tr>"
            )
        p.append("</table>")

    if report.get("cat_freq"):
        p.append(f"<h3 style='{_H3}'>Categorical frequency comparison</h3>")
        for col, cats in report["cat_freq"].items():
            p.append(f"<p style='margin-top:6px;margin-bottom:2px;font-weight:bold;'>{col} ({len(cats)} categories)</p>")
            p.append("<table cellspacing='0' cellpadding='0'>")
            p.append(
                f"<tr><th style='{_TH}'>Category</th>"
                f"<th style='{_TH}'>Real %</th>"
                f"<th style='{_TH}'>Synth %</th>"
                f"<th style='{_TH}'>Diff</th></tr>"
            )
            for c in cats[:50]:
                real_pct = c["real_pct"] or 0.0
                synth_pct = c["synth_pct"] or 0.0
                diff = abs(real_pct - synth_pct)
                diff_style = f"{_TD}text-align:right;color:red;" if diff > 5 else f"{_TD}text-align:right;"
                p.append(
                    f"<tr><td style='{_TD}'>{c['category']}</td>"
                    f"<td style='{_TD}text-align:right;'>{real_pct}</td>"
                    f"<td style='{_TD}text-align:right;'>{synth_pct}</td>"
                    f"<td style='{diff_style}'>{diff:.1f}</td></tr>"
                )
            if len(cats) > 50:
                p.append(
                    f"<tr><td style='{_TD}color:#777;' colspan='4'>"
                    f"… {len(cats) - 50} more categories not shown</td></tr>"
                )
            p.append("</table>")

    # ── Correlation heatmap ───────────────────────────────────────────────────
    if report.get("corr_heatmap_b64"):
        p.append(f"<h2 style='{_H2}'>Correlation Structure — Real vs Synthetic</h2>")
        heatmap_b64 = report["corr_heatmap_b64"]
        # A4 at 96 DPI minus 15 mm margins each side ≈ 681 px usable width.
        # QTextDocument ignores width='100%' for images; use explicit px dimensions.
        _PAGE_W_PX = 681
        try:
            import base64 as _b64, struct as _struct
            raw = _b64.b64decode(heatmap_b64)
            # PNG width/height are at bytes 16-23 of the IHDR chunk
            img_w, img_h = _struct.unpack(">II", raw[16:24])
            if img_w > _PAGE_W_PX:
                img_h = int(img_h * _PAGE_W_PX / img_w)
                img_w = _PAGE_W_PX
        except Exception:
            img_w, img_h = _PAGE_W_PX, 400
        p.append(
            f"<img src='data:image/png;base64,{heatmap_b64}'"
            f" width='{img_w}' height='{img_h}'/>"
        )
        if report.get("mean_corr_diff") is not None:
            p.append(
                f"<p style='text-align:center;color:#555;'>"
                f"Mean absolute correlation difference (upper triangle): "
                f"<b>{report['mean_corr_diff']:.4f}</b></p>"
            )

    p.append("</body></html>")
    return "".join(p)


def export_pdf(report: dict, path: str) -> None:
    """Write the report to a PDF file using QPdfWriter + QTextDocument.

    QPrinter.HighResolution (~1200 DPI) makes QTextDocument lay out text at
    physical printer scale, producing invisibly-small characters. QPdfWriter
    at 96 DPI matches screen resolution so fonts render at readable sizes.
    """
    from PySide6.QtCore import QMarginsF, QSizeF
    from PySide6.QtGui import QPageLayout, QPageSize, QPdfWriter, QTextDocument

    html = _render_pdf_html(report)

    writer = QPdfWriter(path)
    writer.setResolution(96)
    writer.setPageLayout(QPageLayout(
        QPageSize(QPageSize.PageSizeId.A4),
        QPageLayout.Orientation.Portrait,
        QMarginsF(15, 15, 15, 15),
        QPageLayout.Unit.Millimeter,
    ))

    doc = QTextDocument()
    doc.setHtml(html)
    # Tell QTextDocument the usable page area so it paginates correctly.
    paint_rect = writer.pageLayout().paintRectPixels(96)
    doc.setPageSize(QSizeF(paint_rect.size()))
    doc.print_(writer)


# ── Excel ─────────────────────────────────────────────────────────────────────

_PURPLE   = "FF9063CD"
_PUR_DARK = "FF5A3A8E"
_PUR_LITE = "FFEDE8F8"
_GREEN    = "FF61A229"
_RED      = "FFC0392B"
_GREY_HDR = "FFF0F0F0"
_WHITE    = "FFFFFFFF"


def _hdr(ws, row, col, value, bold=True, bg=_GREY_HDR, fg="FF333333"):
    from openpyxl.styles import Font, PatternFill, Alignment
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=fg)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return cell


def _val(ws, row, col, value, align="right", bold=False, fg="FF333333"):
    from openpyxl.styles import Font, Alignment
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=fg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    return cell


def _pass_cell(ws, row, col, passed):
    if passed is None:
        _val(ws, row, col, "—", align="center")
    elif passed:
        _val(ws, row, col, "✓", align="center", bold=True, fg=_GREEN)
    else:
        _val(ws, row, col, "✗", align="center", bold=True, fg=_RED)


def _auto_width(ws, min_w=10, max_w=40):
    from openpyxl.utils import get_column_letter
    for col_cells in ws.columns:
        length = max(
            len(str(c.value or "")) for c in col_cells
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = (
            min(max_w, max(min_w, length + 2))
        )


def export_excel(report: dict, path: str) -> None:
    """Write the quality report to a multi-sheet Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XLImage

    wb = Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    thin = Side(style="thin", color="FFD0C5E8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _apply_border(ws, min_row, max_row, min_col, max_col):
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                ws.cell(r, c).border = border

    # ── Sheet 1: Overview ─────────────────────────────────────────────────────
    ws = wb.create_sheet("Overview")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    title = ws.cell(1, 1, "SLS SynthPop Quality Report")
    title.font = Font(bold=True, size=14, color=_WHITE)
    title.fill = PatternFill("solid", fgColor=_PUR_DARK)
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 28

    for r, (label, value) in enumerate([
        ("Original rows", f"{report['n_orig']:,}"),
        ("Synthetic rows", f"{report['n_synth']:,}"),
        ("Mean correlation difference",
         f"{report['mean_corr_diff']:.4f}" if report.get("mean_corr_diff") is not None else "—"),
    ], start=2):
        ws.cell(r, 1, label).font = Font(bold=True, color=_PUR_DARK)
        ws.cell(r, 2, value).alignment = Alignment(horizontal="right")

    # ── Sheet 2: Numeric Statistics ───────────────────────────────────────────
    ns = report.get("numeric_summary")
    if ns and ns["rows"]:
        ws2 = wb.create_sheet("Numeric Statistics")
        num_vars = ns["num_vars"]

        # Header row 1: Stat | var var var …
        _hdr(ws2, 1, 1, "Statistic", bg=_PUR_DARK, fg=_WHITE)
        for i, col in enumerate(num_vars):
            c = 2 + i * 2
            _hdr(ws2, 1, c, col, bg=_PUR_DARK, fg=_WHITE)
            _hdr(ws2, 1, c + 1, "", bg=_PUR_DARK, fg=_WHITE)
            ws2.merge_cells(
                start_row=1, start_column=c, end_row=1, end_column=c + 1
            )

        # Header row 2: blank | Real Synth Real Synth …
        _hdr(ws2, 2, 1, "", bg=_GREY_HDR)
        for i in range(len(num_vars)):
            c = 2 + i * 2
            _hdr(ws2, 2, c, "Real")
            _hdr(ws2, 2, c + 1, "Synth")

        for r_idx, row in enumerate(ns["rows"], start=3):
            _val(ws2, r_idx, 1, row["stat"], align="left", bold=True)
            for i, col in enumerate(num_vars):
                c = 2 + i * 2
                _val(ws2, r_idx, c, row.get(f"{col}__real"))
                _val(ws2, r_idx, c + 1, row.get(f"{col}__synth"))

        _apply_border(ws2, 1, 2 + len(ns["rows"]), 1, 1 + len(num_vars) * 2)
        _auto_width(ws2)

    # ── Sheet 3: KS Tests ─────────────────────────────────────────────────────
    if report.get("ks_tests"):
        ws3 = wb.create_sheet("KS Tests")
        for c, label in enumerate(["Variable", "KS Statistic", "p-value", "Pass (p≥0.05)"], 1):
            _hdr(ws3, 1, c, label, bg=_PUR_DARK, fg=_WHITE)
        for r, row in enumerate(report["ks_tests"], start=2):
            _val(ws3, r, 1, row["variable"], align="left")
            _val(ws3, r, 2, row["statistic"])
            _val(ws3, r, 3, row["pvalue"])
            _pass_cell(ws3, r, 4, row["pass"])
        _apply_border(ws3, 1, 1 + len(report["ks_tests"]), 1, 4)
        _auto_width(ws3)

    # ── Sheet 4: Chi-squared Tests ────────────────────────────────────────────
    if report.get("chi2_tests"):
        ws4 = wb.create_sheet("Chi-squared Tests")
        for c, label in enumerate(["Variable", "χ² Statistic", "p-value", "Pass (p≥0.05)"], 1):
            _hdr(ws4, 1, c, label, bg=_PUR_DARK, fg=_WHITE)
        for r, row in enumerate(report["chi2_tests"], start=2):
            _val(ws4, r, 1, row["variable"], align="left")
            _val(ws4, r, 2, row["statistic"])
            _val(ws4, r, 3, row["pvalue"])
            _pass_cell(ws4, r, 4, row["pass"])
        _apply_border(ws4, 1, 1 + len(report["chi2_tests"]), 1, 4)
        _auto_width(ws4)

    # ── Sheet 5: Categorical Frequencies ─────────────────────────────────────
    if report.get("cat_freq"):
        ws5 = wb.create_sheet("Categorical Frequencies")
        for c, label in enumerate(["Variable", "Category", "Real %", "Synth %", "Diff"], 1):
            _hdr(ws5, 1, c, label, bg=_PUR_DARK, fg=_WHITE)
        r = 2
        for col, cats in report["cat_freq"].items():
            first = True
            for cat in cats:
                real_pct = cat["real_pct"] or 0.0
                synth_pct = cat["synth_pct"] or 0.0
                diff = round(abs(real_pct - synth_pct), 1)
                _val(ws5, r, 1, col if first else "", align="left", bold=first)
                _val(ws5, r, 2, str(cat["category"]), align="left")
                _val(ws5, r, 3, real_pct)
                _val(ws5, r, 4, synth_pct)
                fg = _RED if diff > 5 else "FF333333"
                _val(ws5, r, 5, diff, fg=fg)
                first = False
                r += 1
        _apply_border(ws5, 1, r - 1, 1, 5)
        _auto_width(ws5)

    # ── Sheet 6: Correlation Heatmap ──────────────────────────────────────────
    if report.get("corr_heatmap_b64"):
        ws6 = wb.create_sheet("Correlation Heatmap")
        img_data = base64.b64decode(report["corr_heatmap_b64"])
        img = XLImage(io.BytesIO(img_data))
        # Scale to fit reasonably within Excel (max ~1400px wide at 96dpi)
        scale = min(1.0, 1400 / (img.width or 1400))
        img.width = int(img.width * scale)
        img.height = int(img.height * scale)
        ws6.add_image(img, "A1")
        if report.get("mean_corr_diff") is not None:
            row_offset = max(40, img.height // 20 + 2)
            ws6.cell(row_offset, 1,
                     f"Mean absolute correlation difference (upper triangle): "
                     f"{report['mean_corr_diff']:.4f}")

    wb.save(path)
